from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pypdf import PdfReader, PdfWriter

import os
import uuid
import pdfplumber
import pandas as pd
import re
import sqlite3
import json
from datetime import datetime
from typing import List, Dict, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
DB_PATH = os.path.join(BASE_DIR, "app.db")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


# =========================================================
# DATABASE
# =========================================================
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS jobs (
            id TEXT PRIMARY KEY,
            original_file_name TEXT,
            safe_file_name TEXT,
            uploaded_at TEXT,
            status TEXT,
            company_name TEXT,
            company_address TEXT,
            company_city TEXT,
            company_country TEXT,
            buyer_name TEXT,
            buyer_address TEXT,
            buyer_city TEXT,
            buyer_country TEXT,
            invoice_number TEXT,
            invoice_date TEXT,
            incoterms TEXT,
            country_export TEXT,
            country_import TEXT,
            port_of_loading TEXT,
            port_of_discharge TEXT,
            transport_mode TEXT,
            total_line_items INTEGER,
            total_quantity TEXT,
            total_value TEXT,
            total_gross_weight TEXT,
            total_net_weight TEXT,
            total_weight TEXT,
            total_number_of_lines INTEGER,
            origin_verified INTEGER,
            weight_matched INTEGER,
            data_sources INTEGER
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS merged_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id TEXT,
            row_index INTEGER,
            row_json TEXT,
            FOREIGN KEY(job_id) REFERENCES jobs(id)
        )
    """)

    conn.commit()
    conn.close()


@app.on_event("startup")
def startup():
    init_db()


@app.get("/")
def home():
    return {"message": "UAE Customs Automation Backend Running"}


# =========================================================
# BASIC HELPERS
# =========================================================
def clean_cell(cell):
    if cell is None:
        return ""
    text = str(cell).replace("\n", " ").replace("\r", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_spaces(text):
    return re.sub(r"\s+", " ", str(text)).strip()


def sanitize_filename(name):
    name = os.path.splitext(name)[0]
    name = re.sub(r'[\\/*?:"<>|]', "", name)
    name = re.sub(r"\s+", "_", name.strip())
    return name


def is_empty_row(row):
    return all(str(cell).strip() == "" for cell in row)


def row_non_empty_count(row):
    return sum(1 for cell in row if str(cell).strip() != "")


def row_text(row):
    return " ".join(str(cell).strip().lower() for cell in row if str(cell).strip() != "")


def parse_number(text):
    if text is None:
        return None
    value = str(text).strip()
    if not value:
        return None

    value = value.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", value)
    if not match:
        return None

    try:
        return float(match.group())
    except Exception:
        return None


def convert_cell_to_number(cell):
    if cell is None:
        return ""

    value = str(cell).strip()
    if value == "":
        return ""

    # Keep long numeric strings as text to avoid Excel scientific notation
    digits_only = re.sub(r"[^\d]", "", value)
    if digits_only and len(digits_only) >= 10:
        return value

    # Keep codes with hyphens/slashes as text
    if re.search(r"[-/]", value):
        return value

    # Keep values starting with 0 as text
    if re.fullmatch(r"0\d+", value):
        return value

    cleaned = value.replace(",", "")
    cleaned = cleaned.replace("$", "")
    cleaned = cleaned.replace("AED", "")
    cleaned = cleaned.replace("USD", "")
    cleaned = cleaned.replace("SAR", "")
    cleaned = cleaned.strip()

    if re.fullmatch(r"\(\d+(\.\d+)?\)", cleaned):
        cleaned = "-" + cleaned[1:-1]

    if re.fullmatch(r"-?\d+", cleaned):
        try:
            return int(cleaned)
        except Exception:
            return value

    if re.fullmatch(r"-?\d+\.\d+", cleaned):
        try:
            return float(cleaned)
        except Exception:
            return value

    return value


def convert_table_numeric_values(table):
    if not table:
        return table

    converted = []
    for row_index, row in enumerate(table):
        if row_index == 0:
            converted.append(row)
            continue
        converted.append([convert_cell_to_number(cell) for cell in row])

    return converted


def build_preview_rows(table, limit=10):
    if not table:
        return []
    preview = []
    for i, row in enumerate(table):
        preview.append([str(cell) for cell in row])
        if i >= limit:
            break
    return preview


# =========================================================
# TABLE EXTRACTION HELPERS
# =========================================================
def group_words_into_rows(words, y_tolerance=6):
    rows = []

    for word in sorted(words, key=lambda w: (w["top"], w["x0"])):
        placed = False

        for row in rows:
            if abs(row["top"] - word["top"]) <= y_tolerance:
                row["words"].append(word)
                row["tops"].append(word["top"])
                row["top"] = sum(row["tops"]) / len(row["tops"])
                placed = True
                break

        if not placed:
            rows.append({
                "top": word["top"],
                "tops": [word["top"]],
                "words": [word]
            })

    return rows


def detect_borderless_columns(words, min_repeats=3, bucket_size=25):
    buckets = {}

    for w in words:
        x = int(w["x0"] // bucket_size) * bucket_size
        buckets[x] = buckets.get(x, 0) + 1

    candidate_x = [x for x, count in buckets.items() if count >= min_repeats]
    candidate_x.sort()

    if len(candidate_x) < 2:
        return []

    merged = [candidate_x[0]]
    for x in candidate_x[1:]:
        if abs(x - merged[-1]) > bucket_size:
            merged.append(x)

    return merged


def assign_words_to_columns(row_words, column_starts):
    row_words = sorted(row_words, key=lambda w: w["x0"])
    cells = [""] * len(column_starts)

    for w in row_words:
        x = w["x0"]
        col_idx = 0

        for i, start in enumerate(column_starts):
            if i == len(column_starts) - 1:
                col_idx = i
                break
            if start <= x < column_starts[i + 1]:
                col_idx = i
                break

        if cells[col_idx]:
            cells[col_idx] += " " + w["text"].strip()
        else:
            cells[col_idx] = w["text"].strip()

    return [c.strip() for c in cells]


def looks_like_borderless_header(row):
    text = " ".join(str(c).lower() for c in row if str(c).strip())
    keywords = ["item", "description", "qty", "quantity", "rate", "price", "amount", "total", "hs", "code"]
    matches = sum(1 for k in keywords if k in text)
    return matches >= 2


def extract_borderless_tables_from_page(page):
    words = page.extract_words(
        x_tolerance=2,
        y_tolerance=2,
        keep_blank_chars=False,
        use_text_flow=True
    )

    if not words:
        return []

    rows = group_words_into_rows(words, y_tolerance=5)
    if not rows:
        return []

    column_starts = detect_borderless_columns(words, min_repeats=3, bucket_size=30)
    if len(column_starts) < 2:
        return []

    rebuilt = []
    for row in rows:
        cells = assign_words_to_columns(row["words"], column_starts)
        if any(str(c).strip() for c in cells):
            rebuilt.append(cells)

    cleaned = []
    for row in rebuilt:
        non_empty = sum(1 for c in row if str(c).strip())
        if non_empty >= 2:
            cleaned.append(row)

    if len(cleaned) < 2:
        return []

    header_found = any(looks_like_borderless_header(r) for r in cleaned[:5])
    if not header_found:
        return []

    return cleaned


def looks_like_item_header(row):
    text = row_text(row)
    header_keywords = [
        "item", "description", "qty", "quantity", "unit", "price", "rate",
        "amount", "total", "value", "hs code", "hs", "origin", "country",
        "gross weight", "net weight", "arabic description"
    ]
    matches = sum(1 for keyword in header_keywords if keyword in text)
    return matches >= 2


def is_total_row(row):
    text = row_text(row)
    total_keywords = [
        "subtotal", "sub total", "grand total", "invoice total",
        "total qty", "total quantity", "total quantities",
        "total gross weight", "total net weight",
        "gross weight total", "net weight total", "net total",
    ]

    if any(keyword in text for keyword in total_keywords):
        return True

    if "total" in text and not looks_like_item_header(row):
        return True

    return False


def is_noise_row(row):
    text = row_text(row)

    noise_keywords = [
        "tel", "telephone", "fax", "email", "www", "website",
        "po box", "p.o. box", "trn", "vat", "customer copy",
        "page ", "bank", "swift", "iban", "beneficiary",
        "signature", "stamp", "authorized signatory",
        "shipper", "consignee"
    ]

    if any(keyword in text for keyword in noise_keywords):
        if not is_total_row(row):
            return True

    return False


def normalize_rows(table):
    cleaned_rows = []

    for row in table:
        row = [clean_cell(cell) for cell in row]

        if is_empty_row(row):
            continue

        if row_non_empty_count(row) < 2 and not is_total_row(row):
            continue

        if is_noise_row(row):
            continue

        cleaned_rows.append(row)

    if not cleaned_rows:
        return []

    max_cols = max(len(row) for row in cleaned_rows)
    normalized = [row + [""] * (max_cols - len(row)) for row in cleaned_rows]

    df = pd.DataFrame(normalized)
    df = df.replace("", pd.NA).dropna(axis=1, how="all").fillna("")

    return df.values.tolist()


def merge_multiline_rows(rows):
    if not rows:
        return rows

    merged = [rows[0]]

    for row in rows[1:]:
        non_empty = row_non_empty_count(row)

        if is_total_row(row):
            merged.append(row)
            continue

        if non_empty <= 2 and merged:
            prev = merged[-1]
            max_len = max(len(prev), len(row))
            prev = prev + [""] * (max_len - len(prev))
            row = row + [""] * (max_len - len(row))

            for i in range(max_len):
                if str(row[i]).strip():
                    if str(prev[i]).strip():
                        prev[i] = f"{prev[i]} {row[i]}".strip()
                    else:
                        prev[i] = row[i]
            merged[-1] = prev
        else:
            merged.append(row)

    return merged


def score_table(table):
    if not table:
        return 0

    rows = len(table)
    cols = max(len(r) for r in table) if table else 0
    filled = sum(1 for row in table for cell in row if str(cell).strip() != "")

    header_bonus = 0
    total_bonus = 0

    for row in table[:5]:
        if looks_like_item_header(row):
            header_bonus += 20

    for row in table:
        if is_total_row(row):
            total_bonus += 5

    return (rows * 4) + (cols * 3) + filled + header_bonus + total_bonus


# =========================================================
# PDF ROTATION NORMALIZATION
# =========================================================
def normalize_rotated_pdf(input_pdf_path, output_pdf_path):
    reader = PdfReader(input_pdf_path)
    writer = PdfWriter()

    changed = False

    for page in reader.pages:
        rotation = page.get("/Rotate", 0)
        rotation = int(rotation) if rotation is not None else 0
        rotation = rotation % 360

        if rotation != 0:
            page.rotate(-rotation)
            changed = True

        writer.add_page(page)

    with open(output_pdf_path, "wb") as f:
        writer.write(f)

    return changed


def get_best_pdf_for_extraction(pdf_path):
    normalized_pdf_path = pdf_path.replace(".pdf", "_upright.pdf")

    try:
        changed = normalize_rotated_pdf(pdf_path, normalized_pdf_path)
        if changed and os.path.exists(normalized_pdf_path):
            return normalized_pdf_path
    except Exception:
        pass

    return pdf_path


def extract_all_tables(pdf_path):
    candidate_tables = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()

            for table in tables:
                cleaned = normalize_rows(table)
                if len(cleaned) >= 2:
                    cleaned = merge_multiline_rows(cleaned)
                    score = score_table(cleaned)

                    if score >= 20:
                        candidate_tables.append({
                            "page": page_num,
                            "table": cleaned,
                            "score": score,
                            "method": "lined"
                        })

            borderless = extract_borderless_tables_from_page(page)
            if borderless:
                cleaned = normalize_rows(borderless)
                if len(cleaned) >= 2:
                    cleaned = merge_multiline_rows(cleaned)
                    score = score_table(cleaned) + 10

                    if score >= 20:
                        candidate_tables.append({
                            "page": page_num,
                            "table": cleaned,
                            "score": score,
                            "method": "borderless"
                        })

    return candidate_tables


def normalize_header_text(row):
    return tuple(re.sub(r"\s+", " ", str(cell).strip().lower()) for cell in row)


def combine_all_tables(candidate_tables):
    if not candidate_tables:
        return []

    candidate_tables = sorted(candidate_tables, key=lambda x: (x["page"], -x["score"]))

    combined_rows = []
    master_header = None
    seen_headers = set()

    for item in candidate_tables:
        table = item["table"]
        if not table:
            continue

        first_row = table[0]

        if looks_like_item_header(first_row):
            header_key = normalize_header_text(first_row)

            if master_header is None:
                master_header = first_row
                combined_rows.append(first_row)
                seen_headers.add(header_key)

            start_index = 1
        else:
            start_index = 0
            if master_header is None:
                master_header = first_row
                combined_rows.append(first_row)
                start_index = 1

        for row in table[start_index:]:
            if looks_like_item_header(row):
                header_key = normalize_header_text(row)
                if header_key in seen_headers:
                    continue
                seen_headers.add(header_key)
                continue

            combined_rows.append(row)

    if not combined_rows:
        return []

    max_cols = max(len(r) for r in combined_rows)
    normalized = [r + [""] * (max_cols - len(r)) for r in combined_rows]

    unique_rows = []
    seen = set()
    for row in normalized:
        key = tuple(str(c).strip() for c in row)
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)

    return unique_rows


# =========================================================
# SMART HEADER + COLUMN REPAIR
# =========================================================
CANONICAL_COLUMNS = [
    "Item No",
    "Description",
    "HS Code",
    "Qty",
    "Unit",
    "Unit Price",
    "Amount",
    "Origin",
    "Gross Weight",
    "Net Weight",
    "Source"
]

HEADER_ALIASES = {
    "Item No": [
        "item", "item no", "item no.", "sr no", "s no", "no", "#", "line", "line no"
    ],
    "Description": [
        "description", "goods description", "item description", "product", "details",
        "arabic description", "commodity", "name"
    ],
    "HS Code": [
        "hs", "hs code", "hscode", "harmonized code", "tariff code", "code"
    ],
    "Qty": [
        "qty", "quantity", "quant", "pcs", "pieces", "piece", "q'ty"
    ],
    "Unit": [
        "unit", "uom", "unit type", "measure", "packing unit"
    ],
    "Unit Price": [
        "unit price", "price", "rate", "price/unit", "unit rate"
    ],
    "Amount": [
        "amount", "value", "line total", "total", "net amount", "item value"
    ],
    "Origin": [
        "origin", "country", "country of origin", "made in"
    ],
    "Gross Weight": [
        "gross weight", "gross wt", "g.w", "gw", "gross"
    ],
    "Net Weight": [
        "net weight", "net wt", "n.w", "nw", "net"
    ],
    "Source": [
        "source"
    ],
}

COMMON_COUNTRIES = {
    "uae", "united arab emirates", "china", "india", "saudi", "saudi arabia",
    "pakistan", "japan", "korea", "south korea", "usa", "u.s.a", "united states",
    "uk", "united kingdom", "germany", "italy", "france", "turkey", "oman", "qatar",
    "kuwait", "bahrain", "malaysia", "thailand", "indonesia", "vietnam"
}


def normalize_header_name(text: str) -> str:
    text = normalize_spaces(text).lower()
    text = text.replace(".", "")
    text = text.replace("_", " ")
    return text


def header_match_score(cell_text: str, canonical: str) -> int:
    text = normalize_header_name(cell_text)
    score = 0

    for alias in HEADER_ALIASES.get(canonical, []):
        alias_norm = normalize_header_name(alias)

        if text == alias_norm:
            score = max(score, 100)
        elif alias_norm in text:
            score = max(score, 70)

    if canonical == "HS Code":
        if "hs" in text and "code" in text:
            score = max(score, 100)

    if canonical == "Gross Weight":
        if "gross" in text:
            score = max(score, 75)

    if canonical == "Net Weight":
        if "net" in text:
            score = max(score, 75)

    if canonical == "Unit Price":
        if "price" in text or "rate" in text:
            score = max(score, 70)

    if canonical == "Amount":
        if "amount" in text or "value" in text or text == "total":
            score = max(score, 70)

    return score


def detect_best_header_row(table: List[List[str]], top_n: int = 6) -> int:
    if not table:
        return 0

    best_idx = 0
    best_score = -1

    for idx, row in enumerate(table[:top_n]):
        row_score = 0
        for cell in row:
            for canonical in CANONICAL_COLUMNS:
                row_score += header_match_score(str(cell), canonical)

        if looks_like_item_header(row):
            row_score += 50

        numeric_cells = sum(1 for c in row if parse_number(c) is not None)
        if numeric_cells >= max(2, len(row) // 2):
            row_score -= 20

        if row_score > best_score:
            best_score = row_score
            best_idx = idx

    return best_idx


def is_hs_code_value(value: str) -> bool:
    value = str(value).strip()
    if not value:
        return False

    compact = value.replace(".", "").replace(" ", "").replace("-", "")
    if re.fullmatch(r"\d{4,12}", compact):
        return True
    return False


def is_country_value(value: str) -> bool:
    raw = str(value).strip()
    if not raw:
        return False

    norm = normalize_header_name(raw)
    if norm in COMMON_COUNTRIES:
        return True

    if re.fullmatch(r"[A-Z]{2}", raw):
        return True

    return False


def is_source_value(value: str) -> bool:
    value = str(value).strip().lower()
    return value in {"ci", "ai", "manual", "ocr", "system"}


def numeric_density(values: List[str]) -> float:
    if not values:
        return 0.0
    cnt = sum(1 for v in values if parse_number(v) is not None)
    return cnt / len(values)


def avg_text_length(values: List[str]) -> float:
    vals = [len(str(v).strip()) for v in values if str(v).strip()]
    return sum(vals) / len(vals) if vals else 0.0


def mostly_integers(values: List[str]) -> float:
    if not values:
        return 0.0
    cnt = 0
    for v in values:
        num = parse_number(v)
        if num is not None and abs(num - round(num)) < 1e-9:
            cnt += 1
    return cnt / len(values)


def count_hs_like(values: List[str]) -> int:
    return sum(1 for v in values if is_hs_code_value(v))


def count_country_like(values: List[str]) -> int:
    return sum(1 for v in values if is_country_value(v))


def count_source_like(values: List[str]) -> int:
    return sum(1 for v in values if is_source_value(v))


def score_column_for_canonical(header_cell: str, values: List[str], canonical: str) -> int:
    score = 0
    header_score = header_match_score(header_cell, canonical)
    score += header_score

    non_empty_values = [str(v).strip() for v in values if str(v).strip()]
    num_density = numeric_density(non_empty_values)
    integer_density = mostly_integers(non_empty_values)
    avg_len = avg_text_length(non_empty_values)
    hs_count = count_hs_like(non_empty_values)
    country_count = count_country_like(non_empty_values)
    source_count = count_source_like(non_empty_values)

    if canonical == "Description":
        if avg_len >= 12:
            score += 35
        if num_density < 0.45:
            score += 15

    elif canonical == "HS Code":
        score += hs_count * 12
        if num_density > 0.5:
            score += 15

    elif canonical == "Qty":
        if integer_density > 0.6:
            score += 30
        if avg_len < 8:
            score += 10

    elif canonical == "Unit":
        common_units = {"pcs", "pc", "kg", "kgs", "set", "sets", "box", "boxes", "ctn", "carton"}
        unit_hits = sum(1 for v in non_empty_values if normalize_header_name(v) in common_units)
        score += unit_hits * 10
        if avg_len <= 5:
            score += 10

    elif canonical == "Unit Price":
        if num_density > 0.6:
            score += 20
        decimal_hits = sum(1 for v in non_empty_values if re.search(r"\d+\.\d{1,4}", str(v)))
        score += decimal_hits * 5

    elif canonical == "Amount":
        if num_density > 0.6:
            score += 20
        larger_hits = sum(1 for v in non_empty_values if (parse_number(v) or 0) >= 10)
        score += min(larger_hits * 3, 20)

    elif canonical == "Origin":
        score += country_count * 12

    elif canonical == "Gross Weight":
        if num_density > 0.6:
            score += 15
        if "gross" in normalize_header_name(header_cell):
            score += 40

    elif canonical == "Net Weight":
        if num_density > 0.6:
            score += 15
        if "net" in normalize_header_name(header_cell):
            score += 40

    elif canonical == "Source":
        score += source_count * 20

    elif canonical == "Item No":
        if integer_density > 0.7:
            score += 20
        if avg_len <= 4:
            score += 10

    return score


def infer_column_mapping(header_row: List[str], body_rows: List[List[str]]) -> Tuple[Dict[int, str], List[int]]:
    if not header_row:
        return {}, []

    max_cols = len(header_row)
    padded_body = []
    for row in body_rows[:20]:
        padded_body.append(row + [""] * (max_cols - len(row)))

    column_scores: Dict[int, Dict[str, int]] = {}

    for col_idx in range(max_cols):
        header_cell = str(header_row[col_idx]) if col_idx < len(header_row) else ""
        samples = [row[col_idx] for row in padded_body if col_idx < len(row)]
        column_scores[col_idx] = {}

        for canonical in CANONICAL_COLUMNS:
            score = score_column_for_canonical(header_cell, samples, canonical)
            column_scores[col_idx][canonical] = score

    assignments: Dict[int, str] = {}
    used_canonicals = set()

    all_candidates = []
    for col_idx, scores in column_scores.items():
        for canonical, score in scores.items():
            all_candidates.append((score, col_idx, canonical))

    all_candidates.sort(reverse=True, key=lambda x: x[0])

    for score, col_idx, canonical in all_candidates:
        if score < 25:
            continue
        if col_idx in assignments:
            continue
        if canonical in used_canonicals and canonical not in {"Description"}:
            continue
        assignments[col_idx] = canonical
        used_canonicals.add(canonical)

    unmatched = [i for i in range(max_cols) if i not in assignments]
    return assignments, unmatched


def standardize_table_columns(table: List[List[str]]) -> List[List[str]]:
    if not table:
        return []

    header_idx = detect_best_header_row(table)
    body = table[header_idx + 1:] if header_idx + 1 < len(table) else []

    if not body:
        body = table[1:] if len(table) > 1 else []

    raw_header = table[header_idx] if table else []
    max_cols = max(len(r) for r in [raw_header] + body) if ([raw_header] + body) else 0

    raw_header = raw_header + [""] * (max_cols - len(raw_header))
    padded_body = [r + [""] * (max_cols - len(r)) for r in body]

    assignments, unmatched = infer_column_mapping(raw_header, padded_body)

    ordered_columns = []
    used_original_cols = set()

    # First add confident canonical columns
    for canonical in CANONICAL_COLUMNS:
        chosen_col = None
        for col_idx, assigned_name in assignments.items():
            if assigned_name == canonical and col_idx not in used_original_cols:
                chosen_col = col_idx
                break
        if chosen_col is not None:
            ordered_columns.append((canonical, chosen_col))
            used_original_cols.add(chosen_col)

    # Keep original PDF headers for unmatched columns
    extra_columns = [idx for idx in range(max_cols) if idx not in used_original_cols]

    for idx in extra_columns:
        col_values = [
            str(row[idx]).strip()
            for row in padded_body
            if idx < len(row) and str(row[idx]).strip()
        ]

        if not col_values:
            continue

        original_header = str(raw_header[idx]).strip()

        if not original_header or original_header.lower() in {
            "nan", "none", "-", "--", "column", "unnamed"
        }:
            original_header = f"Column {idx + 1}"

        existing_names = {name.lower() for name, _ in ordered_columns}
        final_header = original_header
        suffix = 2

        while final_header.lower() in existing_names:
            final_header = f"{original_header} ({suffix})"
            suffix += 1

        ordered_columns.append((final_header, idx))
        used_original_cols.add(idx)

    if not ordered_columns:
        return table

    new_header = [name for name, _ in ordered_columns]
    new_body = []

    for row in padded_body:
        new_row = []
        for _, original_idx in ordered_columns:
            new_row.append(row[original_idx] if original_idx < len(row) else "")
        new_body.append(new_row)

    cleaned_body = []
    for row in new_body:
        if not is_empty_row(row):
            cleaned_body.append(row)

    if not cleaned_body:
        return [new_header]

    return [new_header] + cleaned_body


def add_source_column_if_missing(table):
    if not table:
        return table

    header = table[0]
    normalized = [str(h).strip().lower() for h in header]
    if "source" in normalized:
        return table

    new_table = []
    new_header = header + ["Source"]
    new_table.append(new_header)

    for row in table[1:]:
        new_table.append(row + ["CI"])

    return new_table


def is_summary_row(row):
    text = " ".join(str(cell).strip().lower() for cell in row if str(cell).strip())

    summary_keywords = [
        "subtotal",
        "sub total",
        "invoice total",
        "grand total",
        "total amount",
        "net total",
        "final invoice value",
        "discount",
        "less:",
        "total:"
    ]

    return any(keyword in text for keyword in summary_keywords)


def filter_items_only(table):
    if not table:
        return table

    filtered = []

    for i, row in enumerate(table):
        if i == 0:
            filtered.append(row)
            continue

        if is_summary_row(row):
            continue

        filtered.append(row)

    return filtered


# =========================================================
# SUMMARY + COMPANY INFO
# =========================================================
def find_column_index(header_row, keywords):
    for i, cell in enumerate(header_row):
        cell_text = str(cell).strip().lower()
        for keyword in keywords:
            if keyword in cell_text:
                return i
    return None


def find_last_number_in_row(row):
    nums = [parse_number(c) for c in row if parse_number(c) is not None]
    return nums[-1] if nums else None


def to_display_number(value):
    if value in ("", None):
        return ""
    try:
        num = float(value)
        if num.is_integer():
            return str(int(num))
        return f"{num:.2f}"
    except Exception:
        return str(value)


def build_summary(table):
    summary = {
        "Total Line Items": 0,
        "Total Quantity": "",
        "Total Value": "",
        "Total Gross Weight": "",
        "Total Net Weight": "",
        "Total Number of Lines": 0
    }

    if not table:
        return summary

    summary["Total Number of Lines"] = len(table)

    if len(table) < 2:
        return summary

    header = table[0]
    data_rows = table[1:]

    qty_idx = find_column_index(header, ["qty", "quantity"])
    amount_idx = find_column_index(header, ["amount", "value", "line total", "total"])
    gross_idx = find_column_index(header, ["gross weight", "gross wt", "g.w", "gw"])
    net_idx = find_column_index(header, ["net weight", "net wt", "n.w", "nw"])
    origin_idx = find_column_index(header, ["origin", "country"])
    source_idx = find_column_index(header, ["source"])

    line_items = 0
    calc_qty = 0.0
    calc_value = 0.0
    calc_gross = 0.0
    calc_net = 0.0
    origin_verified = 0
    weight_matched = 0
    data_sources = set()

    explicit_qty_total = None
    explicit_value_total = None
    explicit_gross_total = None
    explicit_net_total = None

    for row in data_rows:
        txt = row_text(row)

        if is_total_row(row):
            last_num = find_last_number_in_row(row)

            if ("qty" in txt or "quantity" in txt) and last_num is not None:
                explicit_qty_total = last_num
            elif "gross" in txt and last_num is not None:
                explicit_gross_total = last_num
            elif "net" in txt and last_num is not None:
                explicit_net_total = last_num
            elif ("total" in txt or "amount" in txt or "value" in txt or "subtotal" in txt) and last_num is not None:
                explicit_value_total = last_num

            continue

        if row_non_empty_count(row) >= 2:
            line_items += 1

        if qty_idx is not None and qty_idx < len(row):
            num = parse_number(row[qty_idx])
            if num is not None:
                calc_qty += num

        if amount_idx is not None and amount_idx < len(row):
            num = parse_number(row[amount_idx])
            if num is not None:
                calc_value += num

        if gross_idx is not None and gross_idx < len(row):
            num = parse_number(row[gross_idx])
            if num is not None:
                calc_gross += num

        if net_idx is not None and net_idx < len(row):
            num = parse_number(row[net_idx])
            if num is not None:
                calc_net += num

        if origin_idx is not None and origin_idx < len(row):
            origin_val = str(row[origin_idx]).strip()
            if origin_val and origin_val != "-":
                origin_verified += 1

        if gross_idx is not None and net_idx is not None and gross_idx < len(row) and net_idx < len(row):
            gross_val = str(row[gross_idx]).strip()
            net_val = str(row[net_idx]).strip()
            if gross_val and gross_val != "-" and net_val and net_val != "-":
                weight_matched += 1

        if source_idx is not None and source_idx < len(row):
            source_val = str(row[source_idx]).strip()
            if source_val:
                data_sources.add(source_val)

    summary["Total Line Items"] = line_items
    summary["Total Quantity"] = to_display_number(
        explicit_qty_total if explicit_qty_total is not None else (calc_qty if calc_qty != 0 else "")
    )
    summary["Total Value"] = to_display_number(
        explicit_value_total if explicit_value_total is not None else (calc_value if calc_value != 0 else "")
    )
    summary["Total Gross Weight"] = to_display_number(
        explicit_gross_total if explicit_gross_total is not None else (calc_gross if calc_gross != 0 else "")
    )
    summary["Total Net Weight"] = to_display_number(
        explicit_net_total if explicit_net_total is not None else (calc_net if calc_net != 0 else "")
    )
    summary["Origin Verified Count"] = origin_verified
    summary["Weight Matched Count"] = weight_matched
    summary["Data Sources Count"] = len(data_sources) if data_sources else 1

    return summary


def extract_text_lines(pdf_path, page_limit=2):
    lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages[:page_limit]:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                line = normalize_spaces(line)
                if line:
                    lines.append(line)
    return lines


def extract_company_info(pdf_path):
    company_info = {
        "Company Name": "",
        "Address": "",
        "Phone": "",
        "Email": "",
        "Invoice Number": "",
        "Invoice Date": "",
        "Buyer Name": "",
        "Buyer Address": "",
        "Buyer City": "",
        "Buyer Country": "",
        "Seller City": "",
        "Seller Country": "",
        "Country of Export": "",
        "Country of Import": "",
        "Port of Loading": "",
        "Port of Discharge": "",
        "Transport Mode": "",
        "Incoterms": ""
    }

    with pdfplumber.open(pdf_path) as pdf:
        first_page_text = pdf.pages[0].extract_text() or ""

    lines = [normalize_spaces(line) for line in first_page_text.split("\n") if normalize_spaces(line)]

    for line in lines[:10]:
        if 3 < len(line) < 100:
            if not re.search(r"invoice|date|address|phone|email|tel|fax|buyer|consignee", line, re.I):
                company_info["Company Name"] = line
                break

    address_lines = []
    for line in lines[:20]:
        lower = line.lower()
        if any(word in lower for word in ["street", "road", "building", "dubai", "sharjah", "abu dhabi", "uae", "po box", "box", "rak", "riyadh"]):
            address_lines.append(line)
    company_info["Address"] = ", ".join(address_lines[:3])

    phone_match = re.search(r"(\+?\d[\d\s\-()]{6,}\d)", first_page_text)
    if phone_match:
        company_info["Phone"] = phone_match.group(1).strip()

    email_match = re.search(r"[\w\.-]+@[\w\.-]+\.\w+", first_page_text)
    if email_match:
        company_info["Email"] = email_match.group(0)

    inv_match = re.search(r"(invoice\s*(no|number)?[:\-\s]*)([A-Za-z0-9\/\-]+)", first_page_text, re.I)
    if inv_match:
        company_info["Invoice Number"] = inv_match.group(3).strip()

    date_match = re.search(r"(\b\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}\b|\b\d{4}[\/\-]\d{1,2}[\/\-]\d{1,2}\b)", first_page_text)
    if date_match:
        company_info["Invoice Date"] = date_match.group(1).strip()

    incoterms_match = re.search(r"\b(EXW|FOB|CIF|CFR|DAP|DDP|FCA|CPT|CIP)\b", first_page_text, re.I)
    if incoterms_match:
        company_info["Incoterms"] = incoterms_match.group(1).upper()

    buyer_match = re.search(r"(buyer|consignee|importer)[:\s]*(.+)", first_page_text, re.I)
    if buyer_match:
        company_info["Buyer Name"] = normalize_spaces(buyer_match.group(2))

    lines_all = extract_text_lines(pdf_path, page_limit=2)
    for line in lines_all:
        low = line.lower()

        if not company_info["Buyer Name"] and any(k in low for k in ["buyer", "consignee", "importer"]):
            cleaned = re.sub(r"^(buyer|consignee|importer)\s*[:\-]?\s*", "", line, flags=re.I).strip()
            if cleaned:
                company_info["Buyer Name"] = cleaned

        if not company_info["Port of Loading"] and ("port of loading" in low or "loading port" in low):
            cleaned = re.sub(r"^(port of loading|loading port)\s*[:\-]?\s*", "", line, flags=re.I).strip()
            company_info["Port of Loading"] = cleaned

        if not company_info["Port of Discharge"] and ("port of discharge" in low or "discharge port" in low):
            cleaned = re.sub(r"^(port of discharge|discharge port)\s*[:\-]?\s*", "", line, flags=re.I).strip()
            company_info["Port of Discharge"] = cleaned

        if not company_info["Transport Mode"] and ("transport mode" in low or "mode of transport" in low):
            cleaned = re.sub(r"^(transport mode|mode of transport)\s*[:\-]?\s*", "", line, flags=re.I).strip()
            company_info["Transport Mode"] = cleaned

        if not company_info["Country of Export"] and ("country of export" in low or "export country" in low):
            cleaned = re.sub(r"^(country of export|export country)\s*[:\-]?\s*", "", line, flags=re.I).strip()
            company_info["Country of Export"] = cleaned

        if not company_info["Country of Import"] and ("country of import" in low or "import country" in low):
            cleaned = re.sub(r"^(country of import|import country)\s*[:\-]?\s*", "", line, flags=re.I).strip()
            company_info["Country of Import"] = cleaned

    upper_address = company_info["Address"].upper()
    if "UAE" in upper_address or "DUBAI" in upper_address or "SHARJAH" in upper_address or "RAK" in upper_address:
        company_info["Seller Country"] = "AE"
    if "RIYADH" in " ".join(lines_all).upper():
        company_info["Buyer City"] = "Riyadh"
        company_info["Buyer Country"] = "SA"

    if not company_info["Country of Export"] and company_info["Seller Country"]:
        company_info["Country of Export"] = company_info["Seller Country"]

    if "DUBAI" in upper_address:
        company_info["Seller City"] = "Dubai"
    elif "SHARJAH" in upper_address:
        company_info["Seller City"] = "Sharjah"
    elif "RAK" in upper_address:
        company_info["Seller City"] = "RAK"

    if not company_info["Country of Import"] and company_info["Buyer Country"]:
        company_info["Country of Import"] = company_info["Buyer Country"]

    return company_info


# =========================================================
# EXCEL FORMATTING
# =========================================================
def should_force_text(header_name, cell_value):
    header = str(header_name).strip().lower()
    value = str(cell_value).strip()

    if not value:
        return False

    text_headers = [
        "hs code", "hs", "item no", "item", "origin", "source", "description"
    ]

    if any(h == header or h in header for h in text_headers):
        return True

    digits_only = re.sub(r"[^\d]", "", value)
    if digits_only and len(digits_only) >= 10:
        return True

    if re.search(r"[-/]", value):
        return True

    return False


def apply_excel_formatting(excel_path):
    wb = load_workbook(excel_path)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sub_header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = left_align

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

        if ws.title == "Items" and ws.max_row >= 2:
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

            for row in ws.iter_rows(min_row=2):
                row_values = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in row]
                row_text_value = " ".join(v for v in row_values if v)

                if "total" in row_text_value or "subtotal" in row_text_value:
                    for cell in row:
                        cell.fill = sub_header_fill
                        cell.font = bold_font

                for idx, cell in enumerate(row):
                    header_name = headers[idx] if idx < len(headers) else ""
                    cell_value = "" if cell.value is None else str(cell.value)

                    if should_force_text(header_name, cell_value):
                        cell.number_format = "@"
                        cell.value = cell_value
        else:
            for row in ws.iter_rows(min_row=2):
                if ws.title in ["Company Info", "Summary"]:
                    if row[0].value:
                        row[0].font = bold_font
                        row[0].fill = sub_header_fill

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in col_cells:
                try:
                    value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 3, 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        ws.freeze_panes = "A2"

    wb.save(excel_path)


# =========================================================
# SAVE JOB
# =========================================================
def save_job(job_id, original_file_name, safe_file_name, payload, merged_items):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("DELETE FROM merged_items WHERE job_id = ?", (job_id,))
    cur.execute("DELETE FROM jobs WHERE id = ?", (job_id,))

    cur.execute("""
        INSERT INTO jobs (
            id, original_file_name, safe_file_name, uploaded_at, status,
            company_name, company_address, company_city, company_country,
            buyer_name, buyer_address, buyer_city, buyer_country,
            invoice_number, invoice_date, incoterms,
            country_export, country_import, port_of_loading, port_of_discharge, transport_mode,
            total_line_items, total_quantity, total_value, total_gross_weight, total_net_weight,
            total_weight, total_number_of_lines, origin_verified, weight_matched, data_sources
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        job_id,
        original_file_name,
        safe_file_name,
        payload["uploaded_at"],
        payload["status"],
        payload["company_name"],
        payload["company_address"],
        payload["company_city"],
        payload["company_country"],
        payload["buyer_name"],
        payload["buyer_address"],
        payload["buyer_city"],
        payload["buyer_country"],
        payload["invoice_number"],
        payload["invoice_date"],
        payload["incoterms"],
        payload["country_export"],
        payload["country_import"],
        payload["port_of_loading"],
        payload["port_of_discharge"],
        payload["transport_mode"],
        payload["total_line_items"],
        payload["total_quantity"],
        payload["total_value"],
        payload["total_gross_weight"],
        payload["total_net_weight"],
        payload["total_weight"],
        payload["total_number_of_lines"],
        payload["origin_verified"],
        payload["weight_matched"],
        payload["data_sources"],
    ))

    for idx, row in enumerate(merged_items):
        cur.execute(
            "INSERT INTO merged_items (job_id, row_index, row_json) VALUES (?, ?, ?)",
            (job_id, idx, json.dumps(row))
        )

    conn.commit()
    conn.close()


# =========================================================
# MAIN ROUTE
# =========================================================
@app.post("/upload-pdf/")
async def upload_pdf(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed.")

    file_id = str(uuid.uuid4())
    safe_original_name = sanitize_filename(file.filename)
    uploaded_at = datetime.now().strftime("%d %b %Y, %I:%M %p")

    pdf_path = os.path.join(UPLOAD_DIR, f"{file_id}.pdf")
    excel_path = os.path.join(OUTPUT_DIR, f"{file_id}.xlsx")

    with open(pdf_path, "wb") as f:
        f.write(await file.read())

    working_pdf_path = get_best_pdf_for_extraction(pdf_path)

    try:
        candidate_tables = extract_all_tables(working_pdf_path)
        if not candidate_tables:
            raise HTTPException(status_code=400, detail="No clean usable tables found in the PDF.")

        combined_table = combine_all_tables(candidate_tables)
        if not combined_table:
            raise HTTPException(status_code=400, detail="Could not combine extracted tables.")

        combined_table = standardize_table_columns(combined_table)
        combined_table = add_source_column_if_missing(combined_table)

        company_info = extract_company_info(working_pdf_path)
        summary = build_summary(combined_table)

        combined_table = filter_items_only(combined_table)
        combined_table = convert_table_numeric_values(combined_table)

        items_df = pd.DataFrame(combined_table)
        company_df = pd.DataFrame(list(company_info.items()), columns=["Field", "Value"])
        summary_df = pd.DataFrame(list(summary.items()), columns=["Field", "Value"])

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            items_df.to_excel(writer, sheet_name="Items", index=False, header=False)
            company_df.to_excel(writer, sheet_name="Company Info", index=False)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

        apply_excel_formatting(excel_path)

        merged_preview = build_preview_rows(combined_table, limit=50)

        response_payload = {
            "id": file_id,
            "message": "PDF processed successfully",
            "excel_file": f"/download-excel/{file_id}?original_name={safe_original_name}",
            "file_name": file.filename,
            "document_code": safe_original_name,
            "uploaded_at": uploaded_at,
            "status": "Processed",
            "company_name": company_info.get("Company Name", ""),
            "company_address": company_info.get("Address", ""),
            "company_city": company_info.get("Seller City", ""),
            "company_country": company_info.get("Seller Country", ""),
            "buyer_name": company_info.get("Buyer Name", ""),
            "buyer_address": company_info.get("Buyer Address", ""),
            "buyer_city": company_info.get("Buyer City", ""),
            "buyer_country": company_info.get("Buyer Country", ""),
            "invoice_number": company_info.get("Invoice Number", ""),
            "invoice_date": company_info.get("Invoice Date", ""),
            "incoterms": company_info.get("Incoterms", ""),
            "country_export": company_info.get("Country of Export", ""),
            "country_import": company_info.get("Country of Import", ""),
            "port_of_loading": company_info.get("Port of Loading", ""),
            "port_of_discharge": company_info.get("Port of Discharge", ""),
            "transport_mode": company_info.get("Transport Mode", ""),
            "total_line_items": summary.get("Total Line Items", 0),
            "total_quantity": summary.get("Total Quantity", ""),
            "total_value": summary.get("Total Value", ""),
            "total_gross_weight": summary.get("Total Gross Weight", ""),
            "total_net_weight": summary.get("Total Net Weight", ""),
            "total_weight": summary.get("Total Gross Weight", "") or summary.get("Total Net Weight", ""),
            "total_number_of_lines": summary.get("Total Number of Lines", 0),
            "origin_verified": summary.get("Origin Verified Count", 0),
            "weight_matched": summary.get("Weight Matched Count", 0),
            "data_sources": summary.get("Data Sources Count", 1),
            "line_items": summary.get("Total Line Items", 0),
            "merged_line_items": merged_preview
        }

        save_job(file_id, file.filename, safe_original_name, response_payload, merged_preview)
        return response_payload

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing PDF: {str(e)}")


# =========================================================
# DASHBOARD + JOB ROUTES
# =========================================================
@app.get("/dashboard-summary")
def dashboard_summary():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) AS cnt FROM jobs")
    total_files = cur.fetchone()["cnt"]

    cur.execute("SELECT COUNT(*) AS cnt FROM jobs WHERE status = 'Processed'")
    successful = cur.fetchone()["cnt"]

    cur.execute("SELECT COUNT(*) AS cnt FROM jobs WHERE status != 'Processed'")
    pending = cur.fetchone()["cnt"]

    cur.execute("SELECT COALESCE(SUM(total_line_items), 0) AS total_items FROM jobs")
    total_items = cur.fetchone()["total_items"]

    cur.execute("""
        SELECT id, original_file_name, uploaded_at, status, total_line_items, total_value
        FROM jobs
        ORDER BY uploaded_at DESC
        LIMIT 10
    """)
    rows = [dict(r) for r in cur.fetchall()]

    conn.close()

    return {
        "processed_today": total_files,
        "successful_extractions": successful,
        "pending_jobs": pending,
        "excel_exports": total_files,
        "total_items": total_items,
        "recent_uploads": rows
    }


@app.get("/jobs")
def list_jobs():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, original_file_name, uploaded_at, status, total_line_items, total_value
        FROM jobs
        ORDER BY rowid DESC
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


@app.get("/jobs/{job_id}")
def get_job(job_id: str):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM jobs WHERE id = ?", (job_id,))
    job = cur.fetchone()
    if not job:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found.")

    cur.execute("SELECT row_json FROM merged_items WHERE job_id = ? ORDER BY row_index ASC", (job_id,))
    rows = cur.fetchall()

    merged_items = [json.loads(r["row_json"]) for r in rows]

    conn.close()

    result = dict(job)
    result["merged_line_items"] = merged_items
    result["excel_file"] = f"/download-excel/{job_id}?original_name={result['safe_file_name']}"
    return result


@app.get("/download-excel/{file_id}")
def download_excel(file_id: str, original_name: str = "file"):
    excel_path = os.path.join(OUTPUT_DIR, f"{file_id}.xlsx")

    if not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail="Excel file not found.")

    download_name = f"{sanitize_filename(original_name)}_excel_extracted.xlsx"

    return FileResponse(
        path=excel_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
